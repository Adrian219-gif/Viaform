from __future__ import annotations

import argparse
import json
import os
import re
import sqlite3
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Iterator, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


BACKEND_DIR = Path(__file__).resolve().parent.parent
DEFAULT_SOURCE_DIR = BACKEND_DIR / "data" / "rankings" / "source"
DEFAULT_OUTPUT_DIR = BACKEND_DIR / "data" / "rankings"
DEFAULT_OVERALL_FILE = DEFAULT_SOURCE_DIR / "qs_2027_overall.xlsx"
DEFAULT_SUBJECT_FILE = DEFAULT_SOURCE_DIR / "qs_2026_by_subjects.xlsx"
DEFAULT_DATABASE = DEFAULT_OUTPUT_DIR / "qs_rankings.sqlite"
DEFAULT_SUBJECTS_JSON = DEFAULT_OUTPUT_DIR / "qs_subjects.json"


@dataclass(frozen=True)
class HeaderColumns:
    row_number: int
    rank: int
    university: int
    country_region: int


@dataclass(frozen=True)
class RankingRecord:
    university: str
    country_region: str
    ranking_system: str
    edition: int
    scope: str
    subject: Optional[str]
    broad_subject: Optional[str]
    rank_display: str
    rank_min: int
    rank_max: Optional[int]
    source_file: str


def clean_text(value: Any) -> str:
    return " ".join(str(value).replace("\u00a0", " ").split()).strip()


def normalized_header(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]+", " ", clean_text(value).upper()).strip()


def find_header(ws: Worksheet, edition: int, scan_rows: int = 30) -> HeaderColumns:
    for row_number, row in enumerate(
        ws.iter_rows(min_row=1, max_row=min(ws.max_row, scan_rows), values_only=True),
        start=1,
    ):
        headers = [normalized_header(value) for value in row]
        rank = next(
            (
                index
                for index, header in enumerate(headers)
                if header == str(edition)
                or header in {"RANK", "CURRENT RANK", f"{edition} RANK"}
            ),
            None,
        )
        university = next(
            (
                index
                for index, header in enumerate(headers)
                if header in {"INSTITUTION", "UNIVERSITY", "NAME", "INSTITUTION NAME"}
            ),
            None,
        )
        country_region = next(
            (
                index
                for index, header in enumerate(headers)
                if "COUNTRY" in header or "TERRITORY" in header
            ),
            None,
        )
        if rank is not None and university is not None and country_region is not None:
            return HeaderColumns(row_number, rank, university, country_region)
    raise ValueError(f"No ranking header found in worksheet {ws.title!r}")


def find_subject_name(ws: Worksheet, header_row: int) -> str:
    candidates: list[str] = []
    for row in ws.iter_rows(min_row=1, max_row=header_row - 1, values_only=True):
        for value in row:
            if not isinstance(value, str):
                continue
            text = clean_text(value)
            if not text or "QS WORLD UNIVERSITY RANKINGS" in text.upper():
                continue
            candidates.append(text)
    if not candidates:
        raise ValueError(f"No official subject name found in worksheet {ws.title!r}")
    return candidates[-1]


def normalize_rank(value: Any) -> tuple[str, int, Optional[int]]:
    if isinstance(value, bool) or value is None:
        raise ValueError(f"Unsupported rank value: {value!r}")

    if isinstance(value, (int, float)):
        if isinstance(value, float) and not value.is_integer():
            raise ValueError(f"Rank must be a whole number: {value!r}")
        rank = int(value)
        return str(rank), rank, rank

    rank_display = clean_text(value)
    comparable = rank_display.replace("–", "-").replace("—", "-")
    comparable = re.sub(r"\s+", "", comparable)

    single = re.fullmatch(r"=?([0-9]+)", comparable)
    if single:
        rank = int(single.group(1))
        return rank_display, rank, rank

    rank_range = re.fullmatch(r"([0-9]+)-([0-9]+)", comparable)
    if rank_range:
        rank_min, rank_max = map(int, rank_range.groups())
        if rank_min > rank_max:
            raise ValueError(f"Invalid descending rank range: {rank_display!r}")
        return rank_display, rank_min, rank_max

    open_ended = re.fullmatch(r"([0-9]+)\+", comparable)
    if open_ended:
        return rank_display, int(open_ended.group(1)), None

    raise ValueError(f"Unsupported rank format: {rank_display!r}")


def iter_records(
    ws: Worksheet,
    header: HeaderColumns,
    edition: int,
    scope: str,
    source_file: str,
    subject: Optional[str] = None,
) -> Iterator[RankingRecord]:
    required_index = max(header.rank, header.university, header.country_region)
    for row_number, row in enumerate(
        ws.iter_rows(min_row=header.row_number + 1, values_only=True),
        start=header.row_number + 1,
    ):
        values = list(row)
        if len(values) <= required_index:
            continue
        university = clean_text(values[header.university]) if values[header.university] is not None else ""
        country_region = clean_text(values[header.country_region]) if values[header.country_region] is not None else ""
        rank_value = values[header.rank]
        if not university and rank_value is None:
            continue
        if not university or not country_region or rank_value is None:
            raise ValueError(
                f"Incomplete ranking row in {ws.title!r} at Excel row {row_number}"
            )
        rank_display, rank_min, rank_max = normalize_rank(rank_value)
        yield RankingRecord(
            university=university,
            country_region=country_region,
            ranking_system="QS",
            edition=edition,
            scope=scope,
            subject=subject,
            broad_subject=None,
            rank_display=rank_display,
            rank_min=rank_min,
            rank_max=rank_max,
            source_file=source_file,
        )


def load_overall(path: Path) -> tuple[list[RankingRecord], dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if len(workbook.worksheets) != 1:
            raise ValueError("Overall workbook must contain exactly one worksheet")
        ws = workbook.worksheets[0]
        header = find_header(ws, 2027)
        records = list(iter_records(ws, header, 2027, "overall", path.name))
        return records, {
            "worksheet": ws.title,
            "header_row": header.row_number,
            "records": len(records),
        }
    finally:
        workbook.close()


def load_subjects(
    path: Path,
) -> tuple[list[RankingRecord], list[dict[str, Any]], list[dict[str, str]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    records: list[RankingRecord] = []
    subjects: list[dict[str, Any]] = []
    skipped: list[dict[str, str]] = []
    try:
        for ws in workbook.worksheets:
            try:
                header = find_header(ws, 2026)
            except ValueError:
                skipped.append({"worksheet": ws.title, "reason": "no ranking table"})
                continue
            subject = find_subject_name(ws, header.row_number)
            sheet_records = list(
                iter_records(
                    ws,
                    header,
                    2026,
                    "subject",
                    path.name,
                    subject=subject,
                )
            )
            if not sheet_records:
                skipped.append({"worksheet": ws.title, "reason": "empty ranking table"})
                continue
            records.extend(sheet_records)
            subjects.append(
                {
                    "subject": subject,
                    "edition": 2026,
                    "worksheet": ws.title,
                    "source_file": path.name,
                    "record_count": len(sheet_records),
                }
            )
    finally:
        workbook.close()
    subjects.sort(key=lambda item: item["subject"].casefold())
    return records, subjects, skipped


def create_database(path: Path, records: Iterable[RankingRecord]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = path.with_suffix(path.suffix + ".tmp")
    temporary_path.unlink(missing_ok=True)
    connection = sqlite3.connect(temporary_path)
    try:
        connection.executescript(
            """
            PRAGMA journal_mode = DELETE;
            CREATE TABLE rankings (
                id INTEGER PRIMARY KEY,
                university TEXT NOT NULL,
                country_region TEXT NOT NULL,
                ranking_system TEXT NOT NULL CHECK (ranking_system = 'QS'),
                edition INTEGER NOT NULL,
                scope TEXT NOT NULL CHECK (scope IN ('overall', 'subject')),
                subject TEXT,
                broad_subject TEXT,
                rank_display TEXT NOT NULL,
                rank_min INTEGER NOT NULL,
                rank_max INTEGER,
                source_file TEXT NOT NULL
            );
            CREATE INDEX idx_rankings_scope ON rankings(scope);
            CREATE INDEX idx_rankings_edition ON rankings(edition);
            CREATE INDEX idx_rankings_subject ON rankings(subject);
            CREATE INDEX idx_rankings_country_region ON rankings(country_region);
            CREATE INDEX idx_rankings_rank_min ON rankings(rank_min);
            CREATE INDEX idx_rankings_lookup
                ON rankings(scope, edition, subject, country_region, rank_min);
            """
        )
        connection.executemany(
            """
            INSERT INTO rankings (
                university, country_region, ranking_system, edition, scope,
                subject, broad_subject, rank_display, rank_min, rank_max, source_file
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                (
                    record.university,
                    record.country_region,
                    record.ranking_system,
                    record.edition,
                    record.scope,
                    record.subject,
                    record.broad_subject,
                    record.rank_display,
                    record.rank_min,
                    record.rank_max,
                    record.source_file,
                )
                for record in records
            ),
        )
        connection.commit()
        integrity = connection.execute("PRAGMA integrity_check").fetchone()[0]
        if integrity != "ok":
            raise RuntimeError(f"SQLite integrity check failed: {integrity}")
    finally:
        connection.close()
    os.replace(temporary_path, path)


def write_subjects_json(
    path: Path,
    source_file: str,
    subjects: list[dict[str, Any]],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = path.with_suffix(path.suffix + ".tmp")
    payload = {
        "ranking_system": "QS",
        "edition": 2026,
        "source_file": source_file,
        "subject_count": len(subjects),
        "subjects": subjects,
    }
    temporary_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    os.replace(temporary_path, path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import official QS Excel rankings into SQLite")
    parser.add_argument("--overall", type=Path, default=DEFAULT_OVERALL_FILE)
    parser.add_argument("--subjects", type=Path, default=DEFAULT_SUBJECT_FILE)
    parser.add_argument("--database", type=Path, default=DEFAULT_DATABASE)
    parser.add_argument("--subjects-json", type=Path, default=DEFAULT_SUBJECTS_JSON)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    for source in (args.overall, args.subjects):
        if not source.is_file():
            raise FileNotFoundError(f"QS source workbook not found: {source}")

    overall_records, overall_meta = load_overall(args.overall)
    subject_records, subjects, skipped = load_subjects(args.subjects)
    create_database(args.database, [*overall_records, *subject_records])
    write_subjects_json(args.subjects_json, args.subjects.name, subjects)

    print(
        json.dumps(
            {
                "database": str(args.database.resolve()),
                "subjects_json": str(args.subjects_json.resolve()),
                "overall_source": args.overall.name,
                "subject_source": args.subjects.name,
                "overall": overall_meta,
                "subject_worksheets": len(subjects),
                "subject_records": len(subject_records),
                "skipped_worksheets": skipped,
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
